import argparse
import logging
import getpass
import ipaddress
import openpyxl
import os
import yaml
import nornir
import tqdm
import glob
import mac_vendor_lookup
from nornir_napalm.plugins.tasks import napalm_get
from nornir_napalm.plugins.tasks import napalm_cli
from openpyxl.worksheet.table import TableStyleInfo, Table
from openpyxl.utils import get_column_letter
import ntc_templates
import sys


def fetch_and_parse_cli_commands(task):
    result_parsed = {}
    error = {}
    cli_results = task.run(task=napalm_cli, commands=['show cdp neighbors', 'show interfaces', 'show interface status', 'show interfaces switchport', 'show inventory'])
    if not task.results.failed:
        for result in cli_results:
            for cli_task in result.result:
                cli_parsed =  ntc_templates.parse.parse_output(platform='cisco_ios', command=cli_task, data=result.result[cli_task])
                if cli_parsed:
                    result_parsed[str(cli_task).replace(' ','_')] = cli_parsed

        return result_parsed

def open_yaml_file(input_file):
    input_text_folder_and_file = input_file
    try:
        logging.info(f'Opening input file {input_text_folder_and_file}')
        input_text_file = open(input_text_folder_and_file)
    except:
        logging.info(f'Input not exist {input_text_folder_and_file}. Exiting')
        sys.exit(-1)
    try:
        logging.info(f'Convert input text file in yaml {input_text_folder_and_file}')
        input_yaml = yaml.safe_load(input_text_file)
    except:
        logging.info(f'Failed to convert{input_text_folder_and_file} in yaml. Exiting')
        sys.exit(-1)

    return input_yaml

def check_and_clean_output_dir(output_dir):

    if not os.path.exists(output_dir):
        os.mkdir(output_dir)
    else:
        files = glob.glob(output_dir+'/*')
        for file in files:
            os.remove(file)

def create_table_style():
    table_style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    return table_style

def create_table(worksheet, table_name, columns, start_position='A1', end_row='2'):
    # table_style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table_style = create_table_style()
    letter_last_column = get_column_letter(len(columns))
    size = str(start_position)+':' + letter_last_column + str(end_row)
    table = Table(displayName=table_name, ref=size)
    table.tableStyleInfo = table_style
    table._initialise_columns()
    if start_position=='A1':
        for column, column_name, column_number in zip(table.tableColumns, columns, range(1,len(columns)+1)):
            column.name = str(column_name)
            current_column_letter = get_column_letter(column_number)
            worksheet.cell(row=1, column=column_number, value=str(column_name))
            # worksheet.column_dimensions[current_column_letter].bestFit = True
            worksheet.column_dimensions[current_column_letter].width = len(column_name)+4


    worksheet.add_table(table)
    return worksheet

def add_data_to_excel_file(workbook, task_name, entries, switch_name):
    columns_names = build_columns_dict(workbook=workbook, worksheet=task_name)
    match entries:
        case dict():
            max_row = int(workbook[task_name].max_row) + 1
            for column in columns_names:
                match column:
                    case 'device':
                        workbook[task_name].cell(row=max_row, column=columns_names['device'], value=switch_name)
                    case _:
                        workbook[task_name].cell(row=max_row, column=columns_names[column], value=entries[column])
        case list():
            for entry in entries:
                max_row = int(workbook[task_name].max_row) + 1
                for column in columns_names:
                    match column:
                        case 'device':
                            workbook[task_name].cell(row=max_row, column=columns_names['device'], value=switch_name)
                        case 'mac_vendor':
                            try:
                                mac_vendor = mac_vendor_lookup.MacLookup().lookup(entry['mac'])
                            except:
                                mac_vendor = 'Not found'
                            workbook[task_name].cell(row=max_row, column=columns_names['mac_vendor'], value=mac_vendor)
                        case 'trunking_vlans':
                            trunk_vlans =' '.join(entry[column])
                            workbook[task_name].cell(row=max_row, column=columns_names['trunking_vlans'], value=trunk_vlans)
                        case _:
                            workbook[task_name].cell(row=max_row, column=columns_names[column], value=entry[column])
    row = workbook[task_name].max_row
    current_table_size = workbook[task_name].tables[task_name].ref
    new_table_size = current_table_size.split(':')[0] + ':' + current_table_size.split(':')[1][0] + str(row)
    workbook[task_name].tables[task_name].ref = new_table_size




    for column in columns_names:
        column_width = 0
        column_letter = openpyxl.utils.get_column_letter(columns_names[column])
        for row in range(1, workbook[task_name].max_row + 1):
            if workbook[task_name].cell(row=row, column=columns_names[column]).value:
                if len(str(workbook[task_name].cell(row=row, column=columns_names[column]).value)) > column_width:
                    column_width = len(workbook[task_name].cell(row=row, column=columns_names[column]).value)
            if workbook[task_name].column_dimensions[column_letter].width < column_width+1:
                workbook[task_name].column_dimensions[column_letter].width = column_width+1







    return workbook

def write_results_to_excel_from_file(spreadsheets_folder, output_excel_file, output_yaml_folder):

    ###check if output folder Exists
    if not os.path.isdir(spreadsheets_folder):
        logging.info(f'Creating output folder {spreadsheets_folder}')
        os.mkdir(spreadsheets_folder)

    output_file = spreadsheets_folder+output_excel_file
    switches = glob.glob(output_yaml_folder+'switches/*.yaml')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for switch in switches:
        switch_yaml = open_yaml_file(switch)
        current_switch_name = switch.split('/')[len(switch.split('/')) - 1].split('.')[0]
        for task in switch_yaml:
            match task:
                case 'get_facts':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'model', 'os_version', 'serial_number' ]
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

                case 'get_mac_address_table':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'interface', 'mac', 'mac_vendor', 'vlan']
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

                case 'show_interfaces':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'interface', 'description', 'link_status', 'protocol_status', 'ip_address', 'speed', 'mtu', 'bandwidth']
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

                case 'show_inventory':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'descr', 'pid', 'sn', 'vid']
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

                case 'show_cdp_neighbors':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'local_interface', 'neighbor', 'neighbor_interface', 'platform']
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

                case 'show_interface_status':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'port', 'name', 'speed', 'vlan', 'status', 'duplex']
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

                case 'show_interfaces_switchport':
                    if task not in wb:
                        ws = wb.create_sheet(title=task)
                        columns = ['device', 'interface', 'mode', 'switchport', 'switchport_monitor', 'switchport_negotiation', 'access_vlan', 'native_vlan', 'trunking_vlans', 'voice_vlan']
                        ws = create_table(worksheet=ws, table_name=task, columns=columns)
                    wb = add_data_to_excel_file(workbook=wb, task_name=task, entries=switch_yaml[task], switch_name=current_switch_name)

    if wb.sheetnames:
        wb.save(output_file)

def format_and_write_switches_results_to_yaml(results, output_yaml_folder):
    if output_yaml_folder.endswith('/') == False:
        output_yaml_folder = output_yaml_folder + '/'
    output_switch_folders = output_yaml_folder+'switches/'

    if not os.path.exists(output_switch_folders):
        os.makedirs(output_switch_folders, exist_ok=True)
    yaml_data = {}
    yaml_error = {}
    for host in results.keys():
        yaml_data.update({ host: {}})
        for result in range(len(results[host])):
            if results[host][result].result != None and results[host][result].failed == False:
                yaml_data[host].update(results[host][result].result)

    for i in yaml_data:
        if yaml_data[i]:
            file = open(output_switch_folders+'/'+i+'.yaml','w')
            try:
                yaml.safe_dump(yaml_data[i], file)
                file.close()
            except:
                print('failed get error to write yaml file ')

    file = open(output_yaml_folder+'get_errors.yaml', 'w')
    if results.failed == True:
        for device in results.failed_hosts:
            yaml_error.update({device: {}})
            for task in results.failed_hosts[device]:
                if task.failed == True:
                    if 'task' not in yaml_error[device]:
                        yaml_error[device].update({'task': []})
                    yaml_error[device]['task'].append(task.name)
        try:
            yaml.safe_dump(yaml_error, file)
            file.close()
        except:
            print('failed get error to write yaml file ')
    file.close()
    return yaml_data
def tasks(task, napalm_get_bar):
    try:
        task.run(task=napalm_get, getters=['get_mac_address_table', 'get_facts'])
        task.run(fetch_and_parse_cli_commands)
        napalm_get_bar.update()
        tqdm.tqdm.write(f"{task.host}: facts gathered")
    except:
        napalm_get_bar.update()
        tqdm.tqdm.write(f"{task.host}: failed")


def build_columns_dict(workbook, worksheet):
    columns_name = {}
    for column in workbook[worksheet].tables[worksheet].tableColumns:
        columns_name.update({column.name: len(columns_name)+1})
    return columns_name

def set_credentials(nr, username=None, password=None, secret=None):
    if not username:
        username = input("Enter username: ")
    if not password:
        password = getpass.getpass()
    if not secret:
        secret = getpass.getpass()

    nr.inventory.defaults.username = username
    nr.inventory.defaults.password = password
    nr.inventory.defaults.connection_options['napalm'] = nornir.core.inventory.ConnectionOptions(
        extras={"optional_args": {"secret": secret}}
    )

    # for host_obj in nornir.inventory.hosts.values():
    #     host_obj.username = username
    #     host_obj.password = password

def open_excel_file(excel_file_path):
    if not os.path.isfile(excel_file_path):
        logging.info(f'Excel file {excel_file_path} not exist. Exiting.')
        exit(-1)

    logging.info('Opening File %s', excel_file_path)
    wb = openpyxl.load_workbook(excel_file_path)
    return wb


def generate_host_yaml(workbook, nornir_inventory_folder):
    yaml_data = {}
    columns_names = build_columns_dict(workbook=workbook, worksheet='devices')
    max_row = int(workbook['devices'].max_row)
    for current_row, row_value in enumerate(workbook['devices'].iter_rows(min_row=2, max_row=max_row), 2):
        # if current_row[columns_names['state']].value != 'ignored':
        if workbook['devices'].cell(row=current_row, column=columns_names['state']).value != 'ignored':
            # current_hostname = str(current_row[columns_names['hostname']].value).strip()
            current_hostname = str(workbook['devices'].cell(row=current_row, column=columns_names['hostname']).value).strip()
            yaml_data.update({current_hostname: {}})
            yaml_data.update({current_hostname: {'data': {}}})
            yaml_data[current_hostname].update({'groups': []})
            # yaml_data[current_hostname].update(
            #     {
            #         'hostname': str(ipaddress.IPv4Interface(current_row[columns_names['mgmt_ip']].value).ip)
            #     }
            # )
            yaml_data[current_hostname].update(
                {
                    'hostname': str(ipaddress.IPv4Interface(workbook['devices'].cell(row=current_row, column=columns_names['mgmt_ip']).value).ip)
                }
            )
            yaml_data[current_hostname]['groups'].append(
                'ios_xe'
            )



    if not os.path.exists(nornir_inventory_folder):
        os.mkdir(nornir_inventory_folder)

    if nornir_inventory_folder.endswith('/') == False:
        nornir_inventory_folder = nornir_inventory_folder + '/'
    file = open(nornir_inventory_folder+'hosts.yaml', 'w')
    yaml.safe_dump(yaml_data, file)
    file.close()

def main():

    logging.basicConfig(level=logging.INFO, format='%(asctime)s:%(levelname)s:%(message)s', handlers=[logging.FileHandler('automation.log'),logging.StreamHandler()])
    parser = argparse.ArgumentParser(description='Python script to get switches configuration')
    #'Folder with Excel Files'
    parser.add_argument('-e', '--excel-folder', dest='spreadsheets_folder', required=False, default='spreadsheets/', help=argparse.SUPPRESS)
    #'Devices file with Excel Files'
    parser.add_argument('-d', '--devices-file', dest='devices_file', required=False, default='devices.xlsx', help=argparse.SUPPRESS)
    #'nornir configuration file'
    parser.add_argument('-n', '--nornir-config', dest='nornir_config', required=False, default='nornir_config.yaml', help=argparse.SUPPRESS)
    #'nornir inventory folder'
    parser.add_argument('-ni', '--nornir-inventory-folder', dest='nornir_inventory_folder', required=False, default='nornir_inventory/', help=argparse.SUPPRESS)
    #'output yaml folder'
    parser.add_argument('-oy', '--output-yaml-folder', dest='output_yaml_folder', required=False, default='results_yaml/', help=argparse.SUPPRESS)
    #'output excel file'
    parser.add_argument('-oe', '--output-excel-file', dest='output_excel_file', required=False, default='results.xlsx', help=argparse.SUPPRESS)
    parser.add_argument('-sg', '--skip-get-information', dest='skip_get_information', required=False, type=bool, choices=[True, False], default=False, help='Do not collect information from devices')


    args = parser.parse_args()

    if args.spreadsheets_folder.endswith('/') == False:
        args.spreadsheets_folder = args.spreadsheets_folder + '/'

    if args.output_yaml_folder.endswith('/') == False:
        args.output_yaml_folder = args.output_yaml_folder + '/'

    device_excel_file = args.spreadsheets_folder+args.devices_file


    if not args.skip_get_information:
        check_and_clean_output_dir(output_dir=args.output_yaml_folder + '/switches')
        wb = open_excel_file(device_excel_file)
        generate_host_yaml(workbook=wb, nornir_inventory_folder=args.nornir_inventory_folder)
        nr = nornir.InitNornir(config_file=args.nornir_config)
        set_credentials(nr=nr)

        with tqdm.tqdm(total=len(nr.inventory.hosts), desc="gathering facts", ) as napalm_bar:
            results = nr.run(task=tasks, napalm_get_bar=napalm_bar)
        results_dict = format_and_write_switches_results_to_yaml(results=results, output_yaml_folder=args.output_yaml_folder)



    write_results_to_excel_from_file(spreadsheets_folder=args.spreadsheets_folder, output_excel_file=args.output_excel_file, output_yaml_folder=args.output_yaml_folder)






if __name__ == '__main__':
    main()